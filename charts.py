import csv
import datetime as DT
import re
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.ticker as mtick
import pandas as pd
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_ru = int((float(self.salary_from) + float(self.salary_to)) / 2) * currency_to_rub[self.salary_currency]
    def getSalaryRu(self):
        return self.salary_ru
class Vacancy:
    def __init__(self, name, salary, area_name, published_at):
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at
class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.csv_filter(file_name)
    @staticmethod
    def clean_string(text):
        text = re.sub(r'<[^>]*>', '', text).replace('\r\n', ' ').strip()
        return re.sub(' +', ' ', text)
    @staticmethod
    def csv_reader(fileName):
        file_csv = csv.reader(open(fileName, encoding='utf_8_sig'))
        try:
            list_data = [x for x in file_csv]
            titles = list_data[0]
            values = [x for x in list_data[1:] if x.count('') == 0 and len(x) == len(titles)]
            return titles, values
        except:
            print('Пустой файл')
            exit()
    @staticmethod
    def csv_filter(file_name):
        titles, values = DataSet.csv_reader(file_name)
        vacList = []
        for value in values:
            vacDic = {}
            for i in range(0, len(value)):
                ans = DataSet.clean_string(value[i])
                vacDic[titles[i]] = ans
            vacList.append(Vacancy(vacDic['name'], Salary(vacDic['salary_from'], vacDic['salary_to'], vacDic['salary_currency']), vacDic['area_name'], vacDic['published_at']))
        return vacList
class InputConnect:
    def __init__(self):
        self.params = InputConnect.get_params()
    @staticmethod
    def get_params():
        filename = input('Введите название файла: ')
        vacancyNames = input('Введите названия профессии: ')
        return filename, vacancyNames
    @staticmethod
    def convert_data(vacancy):
        return int(DT.datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
    @staticmethod
    def create_years(vacList, years):
        for vac in vacList:
            years.add(InputConnect.convert_data(vac))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))
        return years
    @staticmethod
    def create_area_dic(vacList, areaDic):
        for vac in vacList:
            if vac.area_name in areaDic:
                areaDic[vac.area_name].append(vac.salary.getSalaryRu())
            else:
                areaDic[vac.area_name] = [vac.salary.getSalaryRu()]
        return areaDic

    @staticmethod
    def create_vacancies_dic(vacList, vacsDic):
        for vac in vacList:
            if vac.area_name in vacsDic:
                vacsDic[vac.area_name] += 1
            else:
                vacsDic[vac.area_name] = 1
        return vacsDic

    @staticmethod
    def printing_statistical_data(vacList, vacancyNames):
        years = set()
        years = InputConnect.create_years(vacList, years)
        salaryLevelYear = {year: [] for year in years}
        vacYear = {year: 0 for year in years}
        vacSalaryLevelYear = {year: [] for year in years}
        vacCountsYear = {year: 0 for year in years}
        for vac in vacList:
            year = int(InputConnect.convert_data(vac))
            salaryLevelYear[year].append(vac.salary.getSalaryRu())
            vacYear[year] += 1
            for vacancyName in vacancyNames.split(', '):
                if vacancyName in vac.name:
                    vacSalaryLevelYear[year].append(vac.salary.getSalaryRu())
                    vacCountsYear[year] += 1
        salaryLevelYear = {key: int(sum(value)/ len(value)) if len(value) != 0 else 0 for key, value in salaryLevelYear.items()}
        vacSalaryLevelYear = {key: int(sum(value)/ len(value)) if len(value) != 0 else 0 for key, value in vacSalaryLevelYear.items()}
        areaNameDic = {}
        areaNameDic = InputConnect.create_area_dic(vacList, areaNameDic)
        areaNameList = areaNameDic.items()
        areaNameList = [x for x in areaNameList if len(x[1]) / len(vacList) > 0.01]
        areaNameList = sorted(areaNameList, key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        salaryCities = \
            {item[0]: int(sum(item[1]) / len(item[1])) for item in areaNameList[0: min(len(areaNameList), 10)]}
        vacanciesDic = {}
        vacanciesDic = InputConnect.create_vacancies_dic(vacList, vacanciesDic)
        vacanciesCounts = {x: round(y / len(vacList), 4) for x, y in vacanciesDic.items()}
        vacanciesCounts = {k: val for k, val in vacanciesCounts.items() if val >= 0.01}
        vacanciesCities = dict(sorted(vacanciesCounts.items(), key=lambda item: item[1], reverse=True))
        vacanciesCities = dict(list(vacanciesCities.items())[:10])
        return salaryLevelYear, vacYear, vacSalaryLevelYear, vacCountsYear, salaryCities, vacanciesCities

pars = InputConnect()
if pars.params is not None:
    data = DataSet(pars.params[0])

class Report:
    def __init__(self):
            self.columnsData = pars.printing_statistical_data(data.vacancies_objects, pars.params[1])

    @staticmethod
    def as_text(value):
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def generate_excel(columnsData):
        wb = openpyxl.Workbook()
        sheet1 = wb.active
        sheet1.title = "Статистика по годам"
        sheet2 = wb.create_sheet("Статистика по городам")
        thin = Side(border_style="thin", color="000000")
        heads1 = ["Год", "Средняя зарплата", "Средняя зарплата - " + pars.params[1], "Количество вакансий",
                      "Количество вакансий - " + pars.params[1]]
        heads2 = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        for i, head in enumerate(heads1):
            sheet1.cell(row=1, column=i + 1, value=head).font = Font(bold=True)
        for year, value in columnsData[0].items():
            sheet1.append([year, value, columnsData[2][year], columnsData[1][year], columnsData[3][year]])
        for column in sheet1.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet1.column_dimensions[column[0].column_letter].width = length + 1
        for column in sheet1.columns:
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
        for i, head in enumerate(heads2):
            sheet2.cell(row=1, column=i + 1, value=head).font = Font(bold=True)
        for city, value in columnsData[4].items():
            sheet2.append([city, value, ""])
        count = 1
        for city, value in columnsData[5].items():
            count += 1
            sheet2['D' + str(count)] = city
            sheet2['E' + str(count)] = float(columnsData[5][city])
            sheet2['E' + str(count)].number_format = FORMAT_PERCENTAGE_00
        for column in sheet2.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 2
        for column in sheet2.columns:
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
        wb.save('report.xlsx')

    @staticmethod
    def generate_image(columnsData):
        # df = pd.DataFrame({'Количество вакансий': list(columnsData[1].values()),
        #                    'Количество вакансий\n' + 'Frontend-программист': list(columnsData[3].values())},
        #                   index=list(columnsData[0].keys()))
        # ax = df.plot.bar(ylim=(0, 1500), width=0.97, figsize=(26, 10), fontsize=12)
        # ax.grid(axis='y')
        # for p in ax.patches:
        #         ax.annotate(str(p.get_height()), (p.get_x() + p.get_width() / 2.,
        #         (p.get_height() + 5) - p.get_height()), ha='center', va='center', size=10,
        #         fontweight='bold', xytext=(0, 15), textcoords='offset points')
        # ax.figure.savefig('graphVacCountsYearTest.png')
        # x1 = np.arange(len(list(columnsData[0].keys())))
        # width = 0.4
        fig, ax3 = plt.subplots()
        # figsize=(12, 6)
        # ax1.bar(x1 - width, list(columnsData[0].values()), width, label='средняя з/п')
        # ax1.bar(x1, list(columnsData[2].values()), width, label='з/п ' + 'Frontend-программист')
        # ax1.grid(axis='y')
        # y_major_locator = mtick.MultipleLocator(10000)
        # ax1.yaxis.set_major_locator(y_major_locator)
        # ax1.set_xticks(x1 - 0.2, list(columnsData[0].keys()), rotation=90)
        # ax1.legend(loc='upper left', prop={'size': 10})
        # x2 = np.arange(len(list(columnsData[0].keys())))
        # ax2.grid(axis='y')
        # y_major_locator = mtick.MultipleLocator(100)
        # plt.ylim(0, 1400)
        # ax2.yaxis.set_major_locator(y_major_locator)
        # ax2.bar(x2 - width, list(columnsData[1].values()), width, label='Количество вакансий')
        # ax2.bar(x2, list(columnsData[3].values()), width, label='Количество вакансий\n' + 'Frontend-программист')
        # ax2.set_xticks(x2 - 0.2, list(columnsData[0].keys()), rotation=90)
        # ax2.legend(loc='upper left', prop={'size': 10})
        ax3.barh(list([str(city).replace(' ', '\n').replace('-', '-\n') for city in
                        list(reversed(list(columnsData[4].keys())))]),
                    list(reversed(list(columnsData[4].values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=10)
        ax3.xaxis.set_tick_params(labelsize=10)
        ax3.grid(axis='x')
        x_major_locator = mtick.MultipleLocator(20000)
        ax3.xaxis.set_major_locator(x_major_locator)
        # colors = ('blue', 'purple', 'green', 'skyblue', 'orange', 'red', 'peru', 'olive', 'gold', 'yellowgreen', 'teal')
        # other = 1 - sum([rate for rate in columnsData[5].values()])
        # ax4.pie(list(columnsData[5].values()) + [other], labels=list(columnsData[5].keys()) + ['Другие'],
        #         textprops={'fontsize': 10}, colors=colors)
        fig.tight_layout()
        fig.savefig('graphSalaryCities.png')

rep = Report()
#rep.generate_excel(rep.columnsData)
rep.generate_image(rep.columnsData)