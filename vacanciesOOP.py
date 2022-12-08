import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
import csv
import re
import datetime as DT
from prettytable import PrettyTable


dictionary = {
    'name': 'Название',
    'description': 'Описание',
    'key_skills': 'Навыки',
    'experience_id': 'Опыт работы',
    'premium': 'Премиум-вакансия',
    'employer_name': 'Компания',
    'salary_from': 'Оклад',
    'salary_to': 'Верхняя граница вилки оклада',
    'salary_gross': 'Оклад указан до вычета налогов',
    'salary_currency': 'Идентификатор валюты оклада',
    'area_name': 'Название региона',
    'published_at': 'Дата публикации вакансии',
    "noExperience": "Нет опыта",
    "between1And3": "От 1 года до 3 лет",
    "between3And6": "От 3 до 6 лет",
    "moreThan6": "Более 6 лет",
    "AZN": "Манаты",
    "BYR": "Белорусские рубли",
    "EUR": "Евро",
    "GEL": "Грузинский лари",
    "KGS": "Киргизский сом",
    "KZT": "Тенге",
    "RUR": "Рубли",
    "UAH": "Гривны",
    "USD": "Доллары",
    "UZS": "Узбекский сум"
}


def clean_string(text):
    """
    Убирает теги и заменяет символы
    param text (str): строка из файла .csv
    return:
        String: возвращает отчищенную строку
    """
    text = re.sub(r'<[^>]*>', '', text).replace('\r\n', ' ').strip()
    return re.sub(' +', ' ', text)


def csv_reader(file_name):
    """
    Открывает файл .csv и считывает заголовки и записи, при ошибке выходит из файла
    param file_name (str): название файла .csv для обработки
    return:
        при считывании:
            List[str]: заголовки
            List[list[str]]: записи
        при ошибке считывания:
            String: пустой файл
    """
    file_csv = csv.reader(open(file_name, encoding='utf_8_sig'))
    try:
        list_data = [x for x in file_csv]
        titles = list_data[0]
        values = [x for x in list_data[1:] if x.count('') == 0 and len(x) == len(titles)]
        return titles, values
    except:
        print('Пустой файл')
        exit()


def csv_filter(reader, list_naming):
    """
    Фильтрует все значения и составляет лист словарей, ключи: заголовки, значения: записи
    param reader (List[list[str]]): записи
    param list_naming (List[str]): заголовки
    return:
        List[dict]: лист словарей с вакансиями, сформированный из файла .csv, ключи: заголовки, значения: записи
    """
    vacList = []
    for i in range(0, len(reader)):
        for j in range(0, len(reader[i])):
            reader[i][j] = clean_string(reader[i][j])
    for value in reader:
        vacationDictionary = {}
        count = 0
        for field in value:
            vacationDictionary[list_naming[count]] = field
            count += 1
        vacList.append(vacationDictionary)
    return vacList

def formatter(row):
    """
    Форматирует информацию о вакансиях:
        изменяет значения bool на str
        преобразует значение даты в %d.%m.%Y
        изменяет записи по ключу "salary_gross"
        объединяет несколько записей про зарплату в один заголовок
    param row (List[dict]): лист словарей с вакансиями
    return:
        List[dict]: отформатированный лист словарей с вакансиями
    """
    for vacation in row:
        for key, value in vacation.items():
            if vacation[key] == 'False':
                vacation[key] = 'Нет'
            elif vacation[key] == 'True':
                vacation[key] = 'Да'
            try:
                vacation[key] = dictionary[value]
            except:
                pass

        datetime = DT.datetime.strptime(vacation['published_at'][0:10].replace('-', ''), '%Y%m%d').date()
        vacation['published_at'] = datetime.strftime('%d.%m.%Y')
        if vacation['salary_gross'] == 'Да':
            vacation['salary_gross'] = 'Без вычета налогов'
        else:
            vacation['salary_gross'] = 'С вычетом налогов'
        vacation['salary_from'] = f"{'{0:,}'.format(int(float(vacation['salary_from']))).replace(',', ' ')} - " \
                                 f"{'{0:,}'.format(int(float(vacation['salary_to']))).replace(',', ' ')} " \
                                 f"({vacation['salary_currency']}) ({vacation['salary_gross']})"
        del vacation['salary_currency'], vacation['salary_gross'], vacation['salary_to']
    return row


def print_vacancies(data_vacancies, lines, inputFields):
    """
    Выводит в виде таблицы информацию о вакансиях, по опеределённым параметрам: lines, inputFields
    param data_vacancies (List[dict]): лист словарей с вакансиями
    param lines (str):
        диапозон строк, которые нужно вывести
        пример ввода: "20 30" - включая крайние значения, "20" - с 20 строки до конца, "" - все строки
    param inputFields (str):
        диапозон столбцов, которые нужно вывести
        пример ввода: "Название, Опыт работы, Оклад", "" - вывод всех столбцов
    return:
        String: при отсутствии информации о вакансиях выводит пустую строку
    """
    if len(data_vacancies) == 0:
        print('Нет данных')
        return
    table = PrettyTable()
    table.field_names = ['Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия', 'Компания', 'Оклад', 'Название региона', 'Дата публикации вакансии']
    table.max_width = 20
    table.align = 'l'
    table.hrules = 1
    for item in formatter(data_vacancies):
        column = []
        for key, value in item.items():
            if len(value) > 100:
                column.append(value[:100] + '...')
            else:
                column.append(value)
        table.add_row(column)
    table.add_autoindex('№')
    if inputFields != '':
        fieldList = inputFields.split(', ')
        fieldList.append('№')
        if lines.isdigit() == True:
            print(table.get_string(start = int(lines) - 1, fields = fieldList))
        elif len(lines.split()) == 2:
            borders = lines.split()
            print(table.get_string(start = int(borders[0]) - 1, end = int(borders[1]) - 1, fields = fieldList))
        elif lines == '':
            print(table.get_string(fields = fieldList))
    else:
        if lines.isdigit() == True:
            print(table.get_string(start=int(lines) - 1))
        elif len(lines.split()) == 2:
            borders = lines.split()
            print(table.get_string(start=int(borders[0]) - 1, end=int(borders[1]) - 1))
        elif lines == '':
            print(table)

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
class Salary:
    """
    Класс для представления зарплаты
    Attributes:
        salary_from (int): нижняя граница вилки оклада
        salary_to (int): верхняя граница вилки оклада
        salary_currency (str): валюта оклада
    """
    def __init__(self, salary_from, salary_to, salary_currency):
        """
        Инициализирует объект Salary, вычисляет среднюю зарплату из вилки и переводит в рубли, при помощи словаря - currency_to_rub
        param salary_from (str or int or float):  нижняя граница вилки оклада
        param salary_to (str or int or float): верхняя граница вилки оклада
        param salary_currency (str): валюта оклада
        salary_ru (int): средняя зарплата в рублях
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_ru = int((float(self.salary_from) + float(self.salary_to)) / 2) * currency_to_rub[self.salary_currency]
    def getSalaryRu(self):
        """
        Возвращает среднюю зарплату переведённую в рубли
        return:
            Integer: средняя зарплата в рублях
        """
        return self.salary_ru
class Vacancy:
    """
    Класс для представления вакансий
    Attributes:
        name (str): название
        salary (Salary): объект класса Salary
        area_name (str): название города
        published_at (str): дата публикации
    """
    def __init__(self, name, salary, area_name, published_at):
        """
        Инициализирует объект Vacancy
        param name (str): название
        param salary (Salary): объект класса Salary
        param area_name (str): название города
        param published_at (str): дата публикации
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at
class DataSet:
    """
       Класс для представления списка объектов класса Vacancy
       Attributes:
           file_name (str): название файла .csv для обработки
       """
    def __init__(self, file_name):
        """
        Инициализирует объект DataSet, определяет список объектов Vacancy из file_name
        param file_name (str): название файла .csv для обработки
        vacancies_objects (List[Vacancy]): список объектов Vacancy
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.csv_filter(file_name)

    @staticmethod
    def clean_string(text):
        text = re.sub(r'<[^>]*>', '', text).replace('\r\n', ' ').strip()
        return re.sub(' +', ' ', text)

    @staticmethod
    def csv_reader(fileName):
        file_csv = csv.reader(open(fileName, encoding='utf_8_sig'))
        try:
            list_data = [x for x in file_csv]
            titles = list_data[0]
            values = [x for x in list_data[1:] if x.count('') == 0 and len(x) == len(titles)]
            return titles, values
        except:
            print('Пустой файл')
            exit()

    @staticmethod
    def csv_filter(file_name):
        titles, values = DataSet.csv_reader(file_name)
        vacList = []
        for value in values:
            vacDic = {}
            for i in range(0, len(value)):
                ans = DataSet.clean_string(value[i])
                vacDic[titles[i]] = ans
            vacList.append(Vacancy(vacDic['name'], Salary(vacDic['salary_from'], vacDic['salary_to'], vacDic['salary_currency']), vacDic['area_name'], vacDic['published_at']))
        return vacList
class InputConnect:
    def __init__(self):
        self.params = InputConnect.get_params()

    @staticmethod
    def get_params():
        filename = input('Введите название файла: ')
        vacancy = input('Введите название профессии: ')
        return filename, vacancy

    @staticmethod
    def convert_data(vacancy):
        return int(DT.datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))

    @staticmethod
    def create_years(vacList, years):
        for vac in vacList:
            years.add(InputConnect.convert_data(vac))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))
        return years

    @staticmethod
    def create_area_dic(vacList, areaDic):
        for vac in vacList:
            if vac.area_name in areaDic:
                areaDic[vac.area_name].append(vac.salary.getSalaryRu())
            else:
                areaDic[vac.area_name] = [vac.salary.getSalaryRu()]
        return areaDic

    @staticmethod
    def create_vacancies_dic(vacList, vacsDic):
        for vac in vacList:
            if vac.area_name in vacsDic:
                vacsDic[vac.area_name] += 1
            else:
                vacsDic[vac.area_name] = 1
        return vacsDic

    @staticmethod
    def printing_statistical_data(vacList, vacancyName):
        years = set()
        years = InputConnect.create_years(vacList, years)

        salaryLevelYear = {year: [] for year in years}
        vacYear = {year: 0 for year in years}
        vacSalaryLevelYear = {year: [] for year in years}
        vacCountsYear = {year: 0 for year in years}
        for vac in vacList:
            year = int(InputConnect.convert_data(vac))
            salaryLevelYear[year].append(vac.salary.getSalaryRu())
            vacYear[year] += 1
            if vacancyName in vac.name:
                vacSalaryLevelYear[year].append(vac.salary.getSalaryRu())
                vacCountsYear[year] += 1

        salaryLevelYear = {key: int(sum(value)/ len(value)) if len(value) != 0 else 0 for key, value in salaryLevelYear.items()}
        vacSalaryLevelYear = {key: int(sum(value)/ len(value)) if len(value) != 0 else 0 for key, value in vacSalaryLevelYear.items()}

        areaNameDic = {}
        areaNameDic = InputConnect.create_area_dic(vacList, areaNameDic)
        areaNameList = areaNameDic.items()
        areaNameList = [x for x in areaNameList if len(x[1]) / len(vacList) > 0.01]
        areaNameList = sorted(areaNameList, key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        salaryCities = \
            {item[0]: int(sum(item[1]) / len(item[1])) for item in areaNameList[0: min(len(areaNameList), 10)]}
        vacanciesDic = {}
        vacanciesDic = InputConnect.create_vacancies_dic(vacList, vacanciesDic)
        vacanciesCounts = {x: round(y / len(vacList), 4) for x, y in vacanciesDic.items()}
        vacanciesCounts = {k: val for k, val in vacanciesCounts.items() if val >= 0.01}
        vacanciesCities = dict(sorted(vacanciesCounts.items(), key=lambda item: item[1], reverse=True))
        vacanciesCities = dict(list(vacanciesCities.items())[:10])

        print('Динамика уровня зарплат по годам:', salaryLevelYear)
        print('Динамика количества вакансий по годам:', vacYear)
        print('Динамика уровня зарплат по годам для выбранной профессии:', vacSalaryLevelYear)
        print('Динамика количества вакансий по годам для выбранной профессии:', vacCountsYear)
        print('Уровень зарплат по городам (в порядке убывания):', salaryCities)
        print('Доля вакансий по городам (в порядке убывания):', vacanciesCities)
        return salaryLevelYear, vacYear, vacSalaryLevelYear, vacCountsYear, salaryCities, vacanciesCities

class Report:
    def __init__(self):
        self.columnsData = pars.printing_statistical_data(data.vacancies_objects, pars.params[1])

    @staticmethod
    def as_text(value):
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def generate_excel(columnsData):
        wb = openpyxl.Workbook()
        sheet1 = wb.active
        sheet1.title = "Статистика по годам"
        sheet2 = wb.create_sheet("Статистика по городам")
        thin = Side(border_style="thin", color="000000")
        heads1 = ["Год", "Средняя зарплата", "Средняя зарплата - " + pars.params[1], "Количество вакансий", "Количество вакансий - " + pars.params[1]]
        heads2 = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        for i, head in enumerate(heads1):
            sheet1.cell(row=1, column=i + 1, value=head).font = Font(bold=True)
        for year, value in columnsData[0].items():
            sheet1.append([year, value, columnsData[2][year], columnsData[1][year], columnsData[3][year]])
        for column in sheet1.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet1.column_dimensions[column[0].column_letter].width = length + 1
        for column in sheet1.columns:
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
        for i, head in enumerate(heads2):
            sheet2.cell(row=1, column=i + 1, value=head).font = Font(bold=True)
        for city, value in columnsData[4].items():
            sheet2.append([city, value, ""])
        count = 1
        for city, value in columnsData[5].items():
            count += 1
            sheet2['D' + str(count)] = city
            sheet2['E' + str(count)] = float(columnsData[5][city])
            sheet2['E' + str(count)].number_format = FORMAT_PERCENTAGE_00
        for column in sheet2.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 2
        for column in sheet2.columns:
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
        wb.save('report.xlsx')

    @staticmethod
    def generate_image(columnsData):
        x1 = np.arange(len(list(columnsData[0].keys())))
        width = 0.4
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)
        ax1.bar(x1 - width, list(columnsData[0].values()), width, label='средняя з/п')
        ax1.bar(x1, list(columnsData[2].values()), width, label='з/п ' + pars.params[1])
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.set_xticks(x1 - 0.2, list(columnsData[0].keys()), rotation=90)
        ax1.legend(loc='upper left', prop={'size': 8})

        x2 = np.arange(len(list(columnsData[0].keys())))
        ax2.grid(axis='y')
        ax2.bar(x2 - width, list(columnsData[1].values()), width, label='Количество вакансий')
        ax2.bar(x2, list(columnsData[3].values()), width, label='Количество вакансий\n' + pars.params[1])
        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        ax2.set_xticks(x2 - 0.2, list(columnsData[0].keys()), rotation=90)
        ax2.legend(loc='upper left', prop={'size': 8})

        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(city).replace(' ', '\n').replace('-', '-\n') for city in list(reversed(list(columnsData[4].keys())))]),
                list(reversed(list(columnsData[4].values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([rate for rate in columnsData[5].values()])
        ax4.pie(list(columnsData[5].values()) + [other], labels=list(columnsData[5].keys()) + ['Другие'],
                textprops={'fontsize': 6})
        fig.tight_layout()
        fig.savefig('graph.png')

    @staticmethod
    def generate_pdf(columnsData):
        vacCities = columnsData[5]
        vacCities = vacCities.items()
        vacCities = {x[0]: str(f'{x[1] * 100:,.2f}%').replace('.', ',') for x in vacCities}
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        header_year = ["Год", "Средняя зарплата", "Средняя зарплата - " + pars.params[1], "Количество вакансий",
                       "Количество вакансий - " + pars.params[1]]
        header_city = ["Город", "Уровень зарплат", "Город", "Доля вакансий"]
        pdf_template = template.render({'profession': pars.params[1],
                                        'salaryLevelYear': columnsData[0],
                                        'vacYear': columnsData[1],
                                        'vacSalaryLevelYear': columnsData[2],
                                        'vacCountsYear': columnsData[3], 'salaryCities': columnsData[4],
                                        'vacanciesCities': vacCities, 'header_year': header_year,
                                        'header_city': header_city})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

if input('Вакансии или Статистика: ') == 'Вакансии':
    titles, values = csv_reader(input())
    print_vacancies(csv_filter(values, titles), input(), input())
else:
    pars = InputConnect()
    data = DataSet(pars.params[0])
    rep = Report()
    rep.generate_excel(rep.columnsData)
    rep.generate_image(rep.columnsData)
    rep.generate_pdf(rep.columnsData)
